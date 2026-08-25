import openpyxl
from datetime import datetime
from collections import Counter
import re

wb = openpyxl.load_workbook(r'C:\Users\admin\Documents\HiTech files\application\op_history_08_25_0139.xlsx')
ws = wb['History']

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    rows.append({h: ws.cell(r, c).value for c, h in enumerate(headers, 1)})

def txt(v): return str(v).strip() if v is not None and str(v).strip() else ''

print(f"=== TỔNG QUAN ===")
print(f"Tổng số dòng: {len(rows)}")
print(f"\n=== CÁC CỘT TRONG FILE ===")
for i, h in enumerate(headers, 1):
    print(f"{i}. {h}")

# ID
ids = [txt(r['ID']) for r in rows if r['ID']]
id_normalized = [re.sub(r'[^A-Z0-9]', '', txt(r['ID']).upper()) for r in rows]
id_counter = Counter(id_normalized)
print(f"\n=== CỘT ID ===")
print(f"Tổng ID: {len(ids)}")
print(f"ID duy nhất: {len(set(id_normalized))}")
print(f"ID trùng: {sum(1 for c in id_counter.values() if c > 1)} giá trị")
print(f"Dòng trùng: {sum(c-1 for c in id_counter.values() if c > 1)}")

# Công ty (nhà máy)
companies = Counter(txt(r['Công ty']) for r in rows if txt(r['Công ty']))
print(f"\n=== CÔNG TY (Nhà máy) ===")
print(f"Số nhà máy: {len(companies)}")
for k, v in companies.most_common(10):
    print(f"  {k}: {v}")

# Nhà chính
main_houses = Counter(txt(r['Nhà chính']) for r in rows if txt(r['Nhà chính']))
print(f"\n=== NHÀ CHÍNH ===")
print(f"Có giá trị: {sum(main_houses.values())}")
print(f"Rỗng: {len(rows) - sum(main_houses.values())}")
for k, v in main_houses.items():
    print(f"  '{k}': {v}")

# Người tuyển
recruiters = Counter(txt(r['Người tuyển']) for r in rows if txt(r['Người tuyển']))
print(f"\n=== NGƯỜI TUYỂN (Nội bộ) ===")
print(f"Có giá trị: {sum(recruiters.values())}")
print(f"Số người: {len(recruiters)}")
for k, v in recruiters.most_common():
    print(f"  {k}: {v}")

# Vendor (đối tác)
vendors = Counter(txt(r['Vendor']) for r in rows if txt(r['Vendor']))
print(f"\n=== VENDOR (Đối tác) ===")
print(f"Có giá trị: {sum(vendors.values())}")
print(f"Số đối tác: {len(vendors)}")
for k, v in vendors.items():
    print(f"  '{k}': {v}")

# Kiểm tra thiếu thông tin quan trọng
no_recruiter_vendor = sum(1 for r in rows if not txt(r['Người tuyển']) and not txt(r['Vendor']))
no_join_date = sum(1 for r in rows if not r['Ngày vào làm'])
no_factory = sum(1 for r in rows if not txt(r['Công ty']))
no_main_house = sum(1 for r in rows if not txt(r['Nhà chính']))

print(f"\n=== THIẾU THÔNG TIN QUAN TRỌNG ===")
print(f"Thiếu CẢ Người tuyển & Vendor: {no_recruiter_vendor}")
print(f"Thiếu Ngày vào làm: {no_join_date}")
print(f"Thiếu Công ty: {no_factory}")
print(f"Thiếu Nhà chính: {no_main_house}")

# CCCD
cccd_vals = [re.sub(r'\D', '', txt(r['Số CCCD'])) for r in rows]
cccd_len = Counter(len(c) for c in cccd_vals if c)
print(f"\n=== SỐ CCCD ===")
print(f"Rỗng: {sum(1 for c in cccd_vals if not c)}")
print(f"Độ dài: {dict(cccd_len)}")

# Ngày
print(f"\n=== NGÀY ===")
print(f"Ngày sinh có: {sum(1 for r in rows if r['Ngày sinh'])}")
print(f"Ngày vào làm có: {sum(1 for r in rows if r['Ngày vào làm'])}")
print(f"Ngày nghỉ có: {sum(1 for r in rows if r['Ngày nghỉ'])}")

# Mẫu dữ liệu 2 dòng
print(f"\n=== 2 DÒNG MẪU ===")
for i, r in enumerate(rows[:2], 2):
    print(f"\nDòng {i}:")
    print(f"  ID: {r['ID']}")
    print(f"  Họ tên: {r['Họ tên']}")
    print(f"  Công ty: {r['Công ty']}")
    print(f"  Nhà chính: {r['Nhà chính']}")
    print(f"  Người tuyển: {r['Người tuyển']}")
    print(f"  Vendor: {r['Vendor']}")
    print(f"  Ngày vào làm: {r['Ngày vào làm']}")
    print(f"  Ngày nghỉ: {r['Ngày nghỉ']}")

wb.close()
