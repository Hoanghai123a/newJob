import openpyxl
from datetime import datetime
from collections import Counter
import re

wb = openpyxl.load_workbook(r'C:\Users\admin\Documents\HiTech files\application\op_history_08_25_0139.xlsx')
ws = wb['History']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    rows.append({h: ws.cell(r, c).value for c, h in enumerate(headers, 1)})

def txt(v): return str(v).strip() if v is not None and str(v).strip() else ''

# Kiểu dữ liệu ngày
print("=== KIỂU DỮ LIỆU NGÀY (5 dòng đầu) ===")
for r in rows[:3]:
    print("Ngày sinh:", type(r['Ngày sinh']).__name__, repr(r['Ngày sinh']))
    print("Ngày vào làm:", type(r['Ngày vào làm']).__name__, repr(r['Ngày vào làm']))
    print("Ngày nghỉ:", type(r['Ngày nghỉ']).__name__, repr(r['Ngày nghỉ']))
    print("---")

# Ngày sinh / vào làm / nghỉ - tỉ lệ điền
print("\n=== TỈ LỆ ĐIỀN NGÀY ===")
print("Ngày sinh có:", sum(1 for r in rows if r['Ngày sinh']), "/ rỗng:", sum(1 for r in rows if not r['Ngày sinh']))
print("Ngày vào làm có:", sum(1 for r in rows if r['Ngày vào làm']), "/ rỗng:", sum(1 for r in rows if not r['Ngày vào làm']))
print("Ngày nghỉ có:", sum(1 for r in rows if r['Ngày nghỉ']), "/ rỗng:", sum(1 for r in rows if not r['Ngày nghỉ']))

# Ngày sinh bất thường (năm 2010 xuất hiện ở dòng 1)
print("\n=== PHÂN BỐ NĂM SINH ===")
years = Counter(r['Ngày sinh'].year for r in rows if isinstance(r['Ngày sinh'], datetime))
for y in sorted(years):
    if y > 2007 or y < 1960:
        print(f"  {y}: {years[y]}  <-- nghi ngờ")

# Tình trạng
print("\n=== TÌNH TRẠNG (mẫu) ===")
st = Counter(txt(r['Tình trạng']) for r in rows)
for k,v in st.most_common(8): print(f"  '{k}': {v}")

# Địa chỉ
print("\n=== ĐỊA CHỈ ===")
print("Có:", sum(1 for r in rows if txt(r['Địa chỉ'])), "/ rỗng:", sum(1 for r in rows if not txt(r['Địa chỉ'])))

# Ngân hàng
print("\n=== NGÂN HÀNG / TK ===")
print("Ngân hàng có:", sum(1 for r in rows if txt(r['Ngân hàng'])))
print("Số TK có:", sum(1 for r in rows if txt(r['Số tài khoản'])))

# Cột không dùng
print("\n=== CÁC CỘT 'ĐI LÀM' ===")
for c in ['Tên đi làm','Mã nhân viên đi làm','Số CCCD đi làm','Công việc','Lý do nghỉ','Ghi chú']:
    print(f"  {c} có dữ liệu:", sum(1 for r in rows if txt(r[c])))

# Có Nhà chính rỗng nhưng có Vendor?
print("\n=== NHÀ CHÍNH RỖNG ===")
mh_empty_has_vendor = sum(1 for r in rows if not txt(r['Nhà chính']) and txt(r['Vendor']))
mh_empty_no_vendor = sum(1 for r in rows if not txt(r['Nhà chính']) and not txt(r['Vendor']))
print("Nhà chính rỗng + có Vendor:", mh_empty_has_vendor)
print("Nhà chính rỗng + không Vendor:", mh_empty_no_vendor)

# Không có cả người tuyển lẫn vendor
print("\n=== KHÔNG NGƯỜI TUYỂN & KHÔNG VENDOR ===")
for r in rows:
    if not txt(r['Người tuyển']) and not txt(r['Vendor']):
        print("  ID:", r['ID'], "| Họ tên:", r['Họ tên'])

wb.close()
